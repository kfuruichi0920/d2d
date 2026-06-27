"""Excel (.xlsx) 抽出コマンド"""

from __future__ import annotations
import os
import uuid
from typing import Callable


def run(params: dict, progress: Callable[[str], None]) -> dict:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    blob_path: str = params["blob_path"]
    progress(f"Excel 文書を読み込み中: {os.path.basename(blob_path)}")

    wb = openpyxl.load_workbook(blob_path, data_only=True)
    items = []
    structure = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        progress(f"  シート: {sheet_name}")

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        # 空行を除去
        non_empty = [r for r in rows if any(c.strip() for c in r)]

        sheet_uid = str(uuid.uuid4())
        items.append({
            "item_uid": sheet_uid,
            "item_type": "table",
            "sheet_name": sheet_name,
            "rows": non_empty,
            "row_count": len(non_empty),
            "column_count": max((len(r) for r in non_empty), default=0),
        })
        structure.append({"uid": sheet_uid, "title": sheet_name, "children": []})

    progress(f"抽出完了: {len(items)} シート")
    return {"items": items, "structure_json": structure}
